from __future__ import annotations

import csv
import json

from openpyxl import load_workbook

from pdfscan.exporters import COLUMNS, export_csv, export_excel, export_json


def _sample():
    row = dict.fromkeys(COLUMNS, None)
    row.update({"site": "hr", "pdf_url": "https://x/a.pdf", "violations": 3,
                "tagged": 1, "page_count": 5, "offsite": 0})
    return [row]


def test_csv_roundtrip(tmp_path):
    p = export_csv(_sample(), tmp_path / "o.csv")
    with open(p, encoding="utf-8") as fh:
        rows = list(csv.DictReader(fh))
    assert list(rows[0].keys()) == COLUMNS
    assert rows[0]["site"] == "hr"
    assert rows[0]["violations"] == "3"


def test_json_roundtrip(tmp_path):
    p = export_json(_sample(), tmp_path / "o.json")
    data = json.loads(p.read_text(encoding="utf-8"))
    assert data[0]["pdf_url"].endswith("a.pdf")
    assert data[0]["page_count"] == 5


def test_excel_headers_and_rows(tmp_path):
    p = export_excel(_sample(), tmp_path / "o.xlsx")
    ws = load_workbook(p).active
    assert [c.value for c in ws[1]] == COLUMNS
    assert ws.max_row == 2


def test_collect_rows_adds_classification(tmp_path):
    from pdfscan.config import load_settings
    from pdfscan.db import migrate, session
    from pdfscan.db.repositories import PdfRepository, ReportRepository, SiteRepository
    from pdfscan.exporters.base import collect_rows
    from pdfscan.models import DiscoveredPdf, PdfReport, Site, SiteConfig

    assert "classification" in COLUMNS and "classification_reason" in COLUMNS
    settings = load_settings(
        config_path=tmp_path / "missing.yaml",
        overrides={"database": {"path": str(tmp_path / "t.db")}},
    )
    with session(settings.db_path) as conn:
        migrate(conn)
        sid = SiteRepository(conn).upsert(Site(id=None, name="hr", config=SiteConfig(seeds=["u"])))
        pdfs = PdfRepository(conn)
        pid = pdfs.upsert(DiscoveredPdf("https://x/a.pdf", "https://x/", sid))
        pdfs.set_verified(pid, "h", None)
        ReportRepository(conn).upsert(PdfReport(pdf_hash="h", violations=0, tagged=True))
        conn.commit()

    rows = collect_rows(settings, "hr")
    assert rows[0]["classification"] == "good_to_go"
    assert rows[0]["classification_reason"]


def _classified_rows():
    def row(**over):
        base = dict.fromkeys(COLUMNS, None)
        base.update({"site": "hr", "offsite": 0, "failed_checks": 0})
        base.update(over)
        return base

    return [
        row(pdf_url="https://x/good.pdf", violations=0, tagged=1, title_set=1, language_set=1,
            page_count=2, classification="good_to_go", classification_reason="tagged, no counted violations"),
        row(pdf_url="https://x/scan.pdf", violations=12, tagged=1, image_only=1, has_form=1,
            failed_checks=500, page_count=1, classification="needs_manual_remediation",
            classification_reason="image-only / scanned PDF needs OCR"),
        row(pdf_url="https://x/pending.pdf", violations=None, classification="pending"),
    ]


def test_export_html_groups_and_demarcates_status(tmp_path):
    from pdfscan.exporters import export_html, render_html

    doc = render_html(_classified_rows())
    # status sections present
    assert "Needs manual remediation" in doc
    assert "Good to go" in doc
    assert "Pending verification" in doc
    # status colour-coding hooks + flags
    assert 'class="status manual"' in doc
    assert "image-only" in doc and "form" in doc
    # rows rendered with links
    assert "good.pdf" in doc and "scan.pdf" in doc
    # summary tiles
    assert "Needs manual" in doc and "Auto-taggable" in doc

    p = export_html(_classified_rows(), tmp_path / "report.html")
    assert p.exists()
    assert p.read_text(encoding="utf-8").startswith("<!DOCTYPE html>")


def test_export_html_escapes_content(tmp_path):
    from pdfscan.exporters import render_html

    rows = [
        {
            **dict.fromkeys(COLUMNS, None),
            "site": "hr",
            "pdf_url": "https://x/a.pdf?q=<script>",
            "violations": 0,
            "tagged": 1,
            "classification": "good_to_go",
        }
    ]
    doc = render_html(rows)
    assert "<script>" not in doc  # raw tag must be escaped
    assert "&lt;script&gt;" in doc


def test_collect_rows_includes_owner_and_responsible(tmp_path):
    from pdfscan.config import load_settings
    from pdfscan.db import migrate, session
    from pdfscan.db.repositories import (
        PdfRepository,
        PersonRepository,
        ReportRepository,
        SiteOwnerRepository,
        SiteRepository,
    )
    from pdfscan.exporters.base import collect_rows
    from pdfscan.models import (
        DiscoveredPdf,
        PdfReport,
        Person,
        Site,
        SiteConfig,
        SiteOwner,
    )

    assert "owner" in COLUMNS and "responsible" in COLUMNS
    settings = load_settings(
        config_path=tmp_path / "missing.yaml",
        overrides={"database": {"path": str(tmp_path / "t.db")}},
    )
    with session(settings.db_path) as conn:
        migrate(conn)
        sid = SiteRepository(conn).upsert(Site(id=None, name="hr", config=SiteConfig(seeds=["u"])))
        oid = SiteOwnerRepository(conn).upsert(SiteOwner(id=None, key="grp", label="HR Web"))
        SiteRepository(conn).set_owner("hr", oid)
        people = PersonRepository(conn)
        pid = people.upsert(Person(id=None, employee_id="e1", full_name="Ann", email="ann@x"))
        people.add_membership(pid, oid)
        pdfs = PdfRepository(conn)
        fid = pdfs.upsert(DiscoveredPdf("https://x/a.pdf", "https://x/", sid))
        pdfs.set_verified(fid, "h", None)
        ReportRepository(conn).upsert(PdfReport(pdf_hash="h", violations=0, tagged=True))
        conn.commit()

    rows = collect_rows(settings, "hr")
    assert rows[0]["owner"] == "grp"
    assert "ann@x" in (rows[0]["responsible"] or "")
