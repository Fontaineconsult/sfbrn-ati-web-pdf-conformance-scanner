from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

from pdfscan.exporters.base import COLUMNS


def export_excel(rows: list[dict], out_path: str | os.PathLike) -> Path:
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "PDFs"
    ws.append(COLUMNS)
    for row in rows:
        ws.append([_cell(row.get(col)) for col in COLUMNS])

    ws.freeze_panes = "A2"
    for idx, col in enumerate(COLUMNS, start=1):
        width = max(len(col), 14)
        if col in ("pdf_url", "parent_url", "local_path"):
            width = 60
        ws.column_dimensions[get_column_letter(idx)].width = width

    wb.save(out)
    return out


def _cell(value):
    if isinstance(value, bool):
        return value
    if value is None:
        return None
    return value
